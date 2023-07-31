#!/usr/bin/env python
# coding: utf-8

# ## Распределение по папкам общих ОСВ

# In[ ]:


'''
Скрипт для консолидации выгруженных из 1С общих ОСВ, отчётов по проводкам и ведомостей амортизации, и их распределения
по папкам в соответствии с периодом, по которому они сформированы 
'''
# импортируем библиотеки и модули
import numpy as np
import pandas as pd
import os
import math
import shutil
import copy
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

#prevent SettingWithCopyWarning message from appearing
pd.options.mode.chained_assignment = None


# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = os.path.join('H:',
                      '7_ФЭО',
                      '3_Финансы',
                      'Отчетность и расшифровки',
                      'Расшифровка отчетности',
                      'РСБУ',
                      'Исходные данные')
# запишем путь к папке с исходными данными в переменную
source_1c = os.path.join('H:',
                         '7_ФЭО',
                         '3_Финансы',
                         'Отчетность и расшифровки',
                         'Расшифровка отчетности',
                         'РСБУ',
                         'Исходные данные',
                         'Исходные данные из 1С',
                         'Общие ОСВ')


# In[ ]:


# создадим список с названиями всех общих ОСВ в папке с исходными данными с расширением .xlsx
raw_data_name_list = [raw_data for raw_data in os.listdir(source_1c) if raw_data.startswith('Общая ОСВ') and
                      raw_data.endswith('.xlsx')]

print(f'Исходные данные (Общие ОСВ): {raw_data_name_list}')


# In[ ]:


print('Выполнение скрипта "bdr_consolidation" (консолидация БДР)')
print(f'Чтение БДР из "{source}"')

# создадим справочник из всех датафреймов в списке 'raw_data_name_list'
all_raw_data = {raw_data:pd.read_excel(os.path.join(source_1c, f'{raw_data}')) for raw_data in tqdm(raw_data_name_list)}


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Период' в excel файле
period_location_row = [raw_data.apply(lambda row: row.astype(str)
                                                     .str
                                                     .contains('Период')
                                                     .any(), axis=1)
                               .loc[lambda x: x == True].index[0] for raw_data in all_raw_data.values()]


# In[ ]:


# создадим список из названий столбцов, в которых встречается слово 'Период' в excel файле
period_location_column = [raw_data.apply(lambda row: row.astype(str)
                                                        .str
                                                        .contains('Период')
                                                        .any(), axis=0)
                                  .loc[lambda x: x == True].index[0] for raw_data in all_raw_data.values()]


# In[ ]:


# объединим списки 'period_location_row' и 'period_location_column'
# в список кортежей с номером строки и названием столбца
period_location = list(zip(period_location_row, period_location_column))


# In[ ]:


# создадим пустые списки
period_list_start = []
period_list_end = []

# создадим счётчик
x = 0
# при помощи цикла заполним список 'period_list_start' значениями начала периода у каждого файла
for raw_data in all_raw_data.values():
    period_list_start.append(raw_data.loc[period_location[x][0], period_location[x][1]][8:18])
    x += 1

# создадим счётчик
x = 0
# при помощи цикла заполним список 'period_list_end' значениями конца периода у каждого файла
for raw_data in all_raw_data.values():
    period_list_end.append(raw_data.loc[period_location[x][0], period_location[x][1]][-10:])
    x += 1


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Счет' в excel файле в столбце 'Unnamed: 0',
# чтобы в дальнейшем оставить только строки, содержащие таблицу с данными
rows_to_delete_index = []
for raw_data in all_raw_data.values():
    rows_to_delete_index.append(list(raw_data['Unnamed: 0']).index('Счет'))


# In[ ]:


# объединим списки 'raw_data_name_list' и 'rows_to_delete_index'
# в список кортежей с названием файла и номером строки для удаления
rows_to_delete = list(zip(raw_data_name_list, rows_to_delete_index))


# In[ ]:


# при помощи цикла обновим словарь 'all_raw_data', удалив лишние строки из каждого датафрейма
for raw_data in all_raw_data.items():
    for row in rows_to_delete:
        if raw_data[0] == row[0]:
            all_raw_data.update({raw_data[0]: raw_data[1][row[1]+1:]})


# In[ ]:


# при помощи цикла удалим пустые столбцы в каждом датафрейме из словаря 'all_raw_data'
all_raw_data = {raw_data[0]: raw_data[1].dropna(how='all', axis=1) for raw_data in all_raw_data.items()}


# In[ ]:


# при помощи цикла заменим текущие названия столбцов на значения из 1-ой строки
# и удалим 2 лишние строки в каждом датафрейме из словаря 'all_raw_data'
all_raw_data = {raw_data[0]: raw_data[1].rename(columns=raw_data[1].iloc[0])[2:]
                                        .reset_index(drop=True) for raw_data in all_raw_data.items()}


# In[ ]:


# удалим строку с итогами
all_raw_data = {raw_data[0]: raw_data[1].iloc[:-1] for raw_data in all_raw_data.items()}


# In[ ]:


# переименуем столбцы
for row_key, row_value in all_raw_data.items():
    row_value.columns = ['Субконто 1',
                         'Счёт',
                         'Субконто 2',
                         'Сумма нач. ост. Дт',
                         'Сумма нач. ост. Кт',
                         'Сумма оборотов Дт',
                         'Сумма оборотов Кт',
                         'Сумма кон. ост. Дт',
                         'Сумма кон. ост. Кт']


# In[ ]:


# при помощи цикла поменяем местами столбцы
all_raw_data = {raw_data[0]: raw_data[1].iloc[:, [0, 2, 1, 3, 4, 5, 6, 7, 8]] for raw_data in all_raw_data.items()}


# In[ ]:


# создадим новые столбцы 'Является итогом по счёту' и 'Является итогом по Субконто 1'
for row_key, row_value in all_raw_data.items():
    unique_vals = {val: i for i, val in reversed(list(enumerate(x for x in row_value['Счёт'])))}
    del unique_vals[np.nan]
    unique_vals_list_1 = [value for value in unique_vals.values()]
    row_value['Является итогом по счёту'] = row_value.index.values
    row_value['Является итогом по счёту'] = np.where(row_value['Является итогом по счёту'].isin(unique_vals_list_1),
                                                     'Да',
                                                     'Нет')
    
    na_values = list(row_value['Субконто 1'][lambda x: pd.isnull].index)
    unique_vals_list_2 = na_values + unique_vals_list_1
    row_value['Является итогом по Субконто 1'] = row_value.index.values
    row_value['Является итогом по Субконто 1'] = np.where(~(row_value['Является итогом по Субконто 1'].isin(unique_vals_list_2)),
                                                          'Да',
                                                          'Нет')


# In[ ]:


# заполним пропущенные значения в столбцах "Счёт" и "Субконто 1" значениями предыдущей заполненной ячейки по строкам
for row_key, row_value in all_raw_data.items():
    row_value['Счёт'] = row_value['Счёт'].fillna(method='ffill',
                                                 axis=0)
    row_value['Субконто 1'] = row_value['Субконто 1'].fillna(method='ffill',
                                                             axis=0)


# In[ ]:


# сбросим индексы в каждом датафрейме
for row_key, row_value in all_raw_data.items():
    row_value.reset_index(inplace=True, drop=True)


# In[ ]:


# создадим список 'source_file_path' и заполним его значениями путей к исходным файлам
source_file_path = []
for name in raw_data_name_list:
    source_file_path.append(os.path.join(source, name))


# In[ ]:


# объединим списки 'raw_data_name_list' и 'source_file_path'
# в список кортежей с названием файла и путём к нему
source_file = list(zip(raw_data_name_list, source_file_path))


# In[ ]:


'''
Цикл, который перезаписывает каждый файл из 'all_raw_data', чтобы сохранить внесённые ранее изменения в его структуру,
а также сохраняет файл в формате xlsx в первоначальную папку
'''
print(f'Перезапись ОСВ из "{source}"')
for raw_data in tqdm(all_raw_data.items()):
    for file in source_file:
        if raw_data[0] == file[0]:
            raw_data[1].to_excel(file[1], header=True, index=False)


# In[ ]:


# функция для извлечения номера квартала по дате
def quarter_name(date):
    quarters = {'31.03': '1 квартал',
                '30.06': '2 квартал', 
                '30.09': '3 квартал',
                '31.12': '4 квартал'}
    quarter_name = quarters[date]
    return quarter_name


# In[ ]:


# создадим список с расположением новых папок
folder_location = os.path.join(source,
                               period_list_end[0][-4:],
                               f'{period_list_end[0][-4:]}_{quarter_name(period_list_end[0][:-5])}')
folder_location_1c = os.path.join(folder_location, 'Общие ОСВ')
# создаются только те папки, которые до этого не существовали
if not os.path.exists(folder_location):
    os.makedirs(folder_location)
if not os.path.exists(folder_location_1c):
    os.makedirs(folder_location_1c)   


# In[ ]:


# создадим excel файл 'Общая ОСВ' и добавим в него все датафреймы с исходными данными
with pd.ExcelWriter(os.path.join(source, f'Общая ОСВ_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')) as source_file:
    print(f'Создание листов excel из ОСВ')
    for raw_data in tqdm(all_raw_data.items()):
        raw_data[1].to_excel(source_file,
                             sheet_name=raw_data[0][raw_data[0].rfind('_')+1:].replace(' (XLSX).xlsx', ''),
                             header=True,
                             index=False)


# In[ ]:


# создадим путь к итоговому excel файлу
book_path = os.path.join(source, f'Общая ОСВ_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')
# загрузим excel файл для редактирования
book = load_workbook(book_path)


# In[ ]:


columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'] # создадим список столбцов, к которым будем обращаться


# In[ ]:


# создадим таблицы на каждом листе
for sheet in book:
    table = Table(displayName=sheet.title.replace('-','_').replace(' ','_'), ref=f'A1:K{sheet.max_row}') # создадим таблицу
    style = TableStyleInfo(name='TableStyleLight13', # создадим стиль таблицы
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=True)
    table.tableStyleInfo = style # применим стиль
    sheet.add_table(table) # добавим таблицу на лист


# In[ ]:


# цикл для изменения ширины столбцов на каждом листе в excel файле
for sheet in book:
    for column in columns:
        sheet.column_dimensions[column].width = 15
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 35


# In[ ]:


# цикл для изменения формата ячеек на каждом листе в excel файле
for sheet in book:
    for column in columns:
        if column in ['A', 'B', 'C', 'J', 'K']:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[0]
        else:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[3]


# In[ ]:


# цикл для изменения стиля и выравнивания ячеек на каждом листе в excel файле
for sheet in book:
    for column in columns:
        # изменение параметров шрифта
        sheet[f'{column}1'].font = Font(bold=True, 
                                        color='FFFFFFFF')
        # изменение выравнивания
        sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                  vertical='center',
                                                  wrap_text=True)  
    # изменение выравнивания числовых ячеек
    for cell_tuple in sheet[f'D2:I{sheet.max_row}']:
        for cell in cell_tuple:
            cell.alignment = Alignment(horizontal='right',
                                       vertical='top')


# In[ ]:


book.save(book_path) # сохраним excel файл


# In[ ]:


print(f'Перемещение файла "Общая ОСВ" из "{source}" в новую папку')
# скопируем итоговый файл с расширением .xlsx в новую папку
shutil.copy(book_path, folder_location)
# удалим итоговый файл с расширением .xlsx из первоначальной папки
os.remove(book_path)

# цикл, который копирует каждый файл из 'raw_data_list' в новую папку и удаляет его из папки c исходными данными
for name in tqdm(raw_data_name_list):
    # удалим исходные данные с расширением .xlsx из первоначальной папки
    os.remove(os.path.join(source, name))
    shutil.copy(os.path.join(source_1c, name), folder_location_1c)


# ## Распределение по папкам ведомостей амортизации

# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = os.path.join('H:',
                      '7_ФЭО',
                      '3_Финансы',
                      'Отчетность и расшифровки',
                      'Расшифровка отчетности',
                      'РСБУ',
                      'Исходные данные')
# запишем путь к папке с исходными данными в переменную
source_1c = os.path.join('H:',
                         '7_ФЭО',
                         '3_Финансы',
                         'Отчетность и расшифровки',
                         'Расшифровка отчетности',
                         'РСБУ',
                         'Исходные данные',
                         'Исходные данные из 1С',
                         'Ведомости амортизации')


# In[ ]:


print('Выполнение скрипта "bdr_consolidation" (консолидация БДР)')
print(f'Чтение БДР из "{source}"')

# создадим список с названиями всех общих ОСВ в папке с исходными данными с расширением .xlsx
raw_data_name_list_oc = [raw_data for raw_data in os.listdir(source_1c) if raw_data.startswith('Ведомость амортизации ОС') and
                         raw_data.endswith('.xlsx')]

print(f'Исходные данные (Ведомости амортизации ОС): {raw_data_name_list_oc}')


# In[ ]:


# создадим справочник из всех датафреймов в списке 'raw_data_name_list_oc'
all_raw_data_oc = {raw_data:pd.read_excel(os.path.join(source_1c, f'{raw_data}')) for raw_data in tqdm(raw_data_name_list_oc)}


# In[ ]:


# создадим список столбцов, которые нужно сохранить
columns = [0, -1]


# In[ ]:


# оставим только столбцы из списка 'columns'
all_raw_data_oc = {raw_data[0]: raw_data[1].iloc[:, columns] for raw_data in all_raw_data_oc.items()}


# In[ ]:


# временно переименуем столбцы
for row_key, row_value in all_raw_data_oc.items():
    row_value.columns = ['0', '1']


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Основное средство, Инвентарный номер' в excel файле в столбце '0',
# чтобы в дальнейшем оставить только строки, содержащие таблицу с данными
rows_to_delete_index = []
for raw_data in all_raw_data_oc.values():
    rows_to_delete_index.append(list(raw_data['0']).index('Основное средство, Инвентарный номер'))


# In[ ]:


# объединим списки 'raw_data_name_list_oc' и 'rows_to_delete_index'
# в список кортежей с названием файла и номером строки для удаления
rows_to_delete = list(zip(raw_data_name_list_oc, rows_to_delete_index))


# In[ ]:


# при помощи цикла обновим словарь 'all_raw_data_oc', удалив лишние строки из каждого датафрейма и последнюю строку с итогами
for raw_data in all_raw_data_oc.items():
    for row in rows_to_delete:
        if raw_data[0] == row[0]:
            all_raw_data_oc.update({raw_data[0]: raw_data[1][row[1]+1:]})


# In[ ]:


# переименуем столбцы
for row_key, row_value in all_raw_data_oc.items():
    row_value.columns = ['Основное средство, Инвентарный номер',
                         'Остаточная стоимость']


# In[ ]:


# удалим лишние строки из каждого датафрейма в 'all_raw_data_oc'
all_raw_data_oc = {raw_data[0]: raw_data[1].loc[~raw_data[1]['Основное средство, Инвентарный номер'].isin(['01', '03'])]
                   for raw_data in all_raw_data_oc.items()}


# In[ ]:


# создадим список со счетами в 1С
values = ['01.01',
          '01.03',
          '03.01',
          '03.02',
          '03.03']


# In[ ]:


'''
Создадим цикл, в котором если значение в столбце 'Основное средство, Инвентарный номер' принадлежит к множеству values,
то присваиваем значение из этого столбца в столбец 'Счёт',
иначе присваиваем значение 'NaN' в столбец 'Счёт'
'''
for row_key, row_value in all_raw_data_oc.items():
    row_value['Счёт'] = np.where(row_value['Основное средство, Инвентарный номер'].isin(values),
                                 row_value['Основное средство, Инвентарный номер'],
                                 np.nan)


# In[ ]:


# заполним пропущенные значения в столбцах "Счёт" значениями предыдущей заполненной ячейки по строкам
for row_key, row_value in all_raw_data_oc.items():
    row_value['Счёт'] = row_value['Счёт'].fillna(method='ffill',
                                                 axis=0)


# In[ ]:


'''
Обновим словарь all_raw_data_oc, где ключами будут первые элементы из all_raw_data_oc.items(),
а значениями будут фильтрованные данные из raw_data, где столбец 'Основное средство, Инвентарный номер'
не содержит значения из списка values
'''
all_raw_data_oc = {raw_data[0]: raw_data[1].loc[~raw_data[1]['Основное средство, Инвентарный номер'].isin(values)]
                   for raw_data in all_raw_data_oc.items()}


# In[ ]:


# при помощи цикла поменяем местами столбцы
all_raw_data_oc = {raw_data[0]: raw_data[1].iloc[:, [0, 2, 1]] for raw_data in all_raw_data_oc.items()}


# In[ ]:


# присвоим пустую строку последнему элементу столбца 'Счёт'
for row_key, row_value in all_raw_data_oc.items():
    row_value['Счёт'].iloc[-1:] = ''


# In[ ]:


# сбросим индексы в каждом датафрейме
for row_key, row_value in all_raw_data_oc.items():
    row_value.reset_index(inplace=True, drop=True)


# In[ ]:


# создадим список 'source_file_path' и заполним его значениями путей к исходным файлам
source_file_path = []
for name in raw_data_name_list_oc:
    source_file_path.append(os.path.join(source, name))


# In[ ]:


# объединим списки 'raw_data_name_list_oc' и 'source_file_path'
# в список кортежей с названием файла и путём к нему
source_file = list(zip(raw_data_name_list_oc, source_file_path))


# In[ ]:


'''
Цикл, который перезаписывает каждый файл из 'all_raw_data_oc', чтобы сохранить внесённые ранее изменения в его структуру,
а также сохраняет файл в формате xlsx в первоначальную папку
'''
print(f'Перезапись Ведомости амортизации ОС из "{source}"')
for raw_data in tqdm(all_raw_data_oc.items()):
    for file in source_file:
        if raw_data[0] == file[0]:
            raw_data[1].to_excel(file[1], header=True, index=False)


# In[ ]:


# создадим путь к папке 'Ведомости амортизации'
folder_location_1c = os.path.join(folder_location, 'Ведомости амортизации')
# создаются только те папки, которые до этого не существовали
if not os.path.exists(folder_location_1c):
    os.makedirs(folder_location_1c)   


# In[ ]:


# создадим excel файл 'Ведомость амортизации ОС' и добавим в него все датафреймы с исходными данными
with pd.ExcelWriter(os.path.join(source, f'Ведомость амортизации ОС_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')) as source_file:
    print(f'Создание листов excel из Ведомостей амортизации ОС')
    for raw_data in tqdm(all_raw_data_oc.items()):
        raw_data[1].to_excel(source_file,
                             sheet_name=raw_data[0][raw_data[0].rfind('_')+1:].replace('.xlsx', ''),
                             header=True,
                             index=False)


# In[ ]:


# создадим путь к итоговому excel файлу
book_path_oc = os.path.join(source, f'Ведомость амортизации ОС_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')
# загрузим excel файл для редактирования
book_oc = load_workbook(book_path_oc)


# In[ ]:


columns = ['A', 'B', 'C'] # создадим список столбцов, к которым будем обращаться


# In[ ]:


# создадим таблицы на каждом листе
for sheet in book_oc:
    table = Table(displayName=sheet.title.replace('-','_').replace(' ','_'), ref=f'A1:C{sheet.max_row}') # создадим таблицу
    style = TableStyleInfo(name='TableStyleLight13', # создадим стиль таблицы
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=True)
    table.tableStyleInfo = style # применим стиль
    sheet.add_table(table) # добавим таблицу на лист


# In[ ]:


# цикл для изменения ширины столбцов на каждом листе в excel файле
for sheet in book_oc:
    for column in columns:
        sheet.column_dimensions[column].width = 35


# In[ ]:


# цикл для изменения формата ячеек на каждом листе в excel файле
for sheet in book_oc:
    for column in columns:
        if column in ['A', 'B']:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[0]
        else:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[3]


# In[ ]:


# цикл для изменения стиля и выравнивания ячеек на каждом листе в excel файле
for sheet in book_oc:
    for column in columns:
        # изменение параметров шрифта
        sheet[f'{column}1'].font = Font(bold=True, 
                                        color='FFFFFFFF')
        # изменение выравнивания
        sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                  vertical='center',
                                                  wrap_text=True)  
    # изменение выравнивания числовых ячеек
    for cell_tuple in sheet[f'C2:C{sheet.max_row}']:
        for cell in cell_tuple:
            cell.alignment = Alignment(horizontal='right',
                                       vertical='top')


# In[ ]:


book_oc.save(book_path_oc) # сохраним excel файл


# In[ ]:


print(f'Перемещение файла "Ведомость амортизации" из "{source}" в новую папку')
# скопируем итоговый файл с расширением .xlsx в новую папку
shutil.copy(book_path_oc, folder_location)
# удалим итоговый файл с расширением .xlsx из первоначальной папки
os.remove(book_path_oc)

# цикл, который копирует каждый файл из 'raw_data_list' в новую папку и удаляет его из папки c исходными данными
for name in tqdm(raw_data_name_list_oc):
    # удалим исходные данные с расширением .xlsx из первоначальной папки
    os.remove(os.path.join(source, name))
    shutil.copy(os.path.join(source_1c, name), folder_location_1c)


# ## Распределение по папкам отчётов по проводкам

# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = os.path.join('H:',
                      '7_ФЭО',
                      '3_Финансы',
                      'Отчетность и расшифровки',
                      'Расшифровка отчетности',
                      'РСБУ',
                      'Исходные данные')
# запишем путь к папке с исходными данными в переменную
source_1c = os.path.join('H:',
                         '7_ФЭО',
                         '3_Финансы',
                         'Отчетность и расшифровки',
                         'Расшифровка отчетности',
                         'РСБУ',
                         'Исходные данные',
                         'Исходные данные из 1С',
                         'Отчёты по проводкам')


# In[ ]:


print('Выполнение скрипта "bdr_consolidation" (консолидация БДР)')
print(f'Чтение БДР из "{source}"')

# создадим список с названиями всех общих ОСВ в папке с исходными данными с расширением .xlsx
raw_data_name_list_pr = [raw_data for raw_data in os.listdir(source_1c) if raw_data.startswith('Отчёт по проводкам') and
                         raw_data.endswith('.xlsx')]

print(f'Исходные данные (Ведомости амортизации ОС): {raw_data_name_list_pr}')


# In[ ]:


# создадим справочник из всех датафреймов в списке 'raw_data_name_list_pr'
all_raw_data_pr = {raw_data:pd.read_excel(os.path.join(source_1c, f'{raw_data}')) for raw_data in tqdm(raw_data_name_list_pr)}


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Дебет' в excel файле
debet_location_row = [raw_data.apply(lambda row: row.astype(str)
                                     .eq('Дебет')
                                     .any(), axis=1)
                      .loc[lambda x: x == True].index[0] for raw_data in all_raw_data_pr.values()]


# In[ ]:


# создадим список из названий столбцов, в которых встречается слово 'Дебет' в excel файле
debet_location_column = {raw_data[0]: raw_data[1].apply(lambda row: row.astype(str)
                                                        .eq('Дебет')
                                                        .any(), axis=0)
                         .loc[lambda x: x == True].index[0] for raw_data in all_raw_data_pr.items()}


# In[ ]:


# создадим список из названий столбцов, в которых встречается слово 'Дебет' в excel файле
debet_location_column = {column[0]: all_raw_data_pr[column[0]].columns.get_loc(column[1]) for column in debet_location_column.items()}
debet_location_column = [all_raw_data_pr[column[0]].columns[column[1]+1] for column in debet_location_column.items()]


# In[ ]:


# объединим списки 'debet_location_row' и 'debet_location_column'
# в список кортежей с номером строки и названием столбца
debet_location = list(zip(debet_location_row, debet_location_column))


# In[ ]:


# объединим списки 'raw_data_name_list_pr' и 'debet_location'
# в список кортежей с названием файла и номером строки для удаления
cells_to_replace = dict(zip(raw_data_name_list_pr, debet_location))


# In[ ]:


'''
Если название датафрейма равно имени из cells_to_replace,
то заменим значение ячеек на 'Дебет'
'''
for raw_data in all_raw_data_pr.items():
    for name, cell in cells_to_replace.items():
        if raw_data[0] == name:
            raw_data[1].at[cell[0], cell[1]] = 'Дебет'


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Кредит' в excel файле
credit_location_row = [raw_data.apply(lambda row: row.astype(str)
                                      .eq('Кредит')
                                      .any(), axis=1)
                       .loc[lambda x: x == True].index[0] for raw_data in all_raw_data_pr.values()]


# In[ ]:


# создадим список из названий столбцов, в которых встречается слово 'Кредит' в excel файле
credit_location_column = {raw_data[0]: raw_data[1].apply(lambda row: row.astype(str)
                                                         .eq('Кредит')
                                                         .any(), axis=0)
                          .loc[lambda x: x == True].index[0] for raw_data in all_raw_data_pr.items()}


# In[ ]:


# добавим новый столбец с пустыми значениями в каждый датафрейм, чтобы избежать ошибок
all_raw_data_pr = {raw_data[0]: raw_data[1].assign(temp=[float('nan')] * len(raw_data[1])) for raw_data in all_raw_data_pr.items()}


# In[ ]:


# создадим список из названий столбцов, в которых встречается слово 'Кредит' в excel файле
credit_location_column = {column[0]: all_raw_data_pr[column[0]].columns.get_loc(column[1]) for column in credit_location_column.items()}
credit_location_column = [all_raw_data_pr[column[0]].columns[column[1]+1] for column in credit_location_column.items()]


# In[ ]:


# объединим списки 'credit_location_row' и 'credit_location_column'
# в список кортежей с номером строки и названием столбца
credit_location = list(zip(credit_location_row, credit_location_column))


# In[ ]:


# объединим списки 'raw_data_name_list_pr' и 'credit_location'
# в список кортежей с названием файла и номером строки для удаления
cells_to_replace = dict(zip(raw_data_name_list_pr, credit_location))


# In[ ]:


'''
Если название датафрейма равно имени из cells_to_replace,
то заменим значение ячеек на 'Кредит'
'''
for raw_data in all_raw_data_pr.items():
    for name, cell in cells_to_replace.items():
        if raw_data[0] == name:
            raw_data[1].at[cell[0], cell[1]] = 'Кредит'


# In[ ]:


# при помощи цикла удалим пустые столбцы в каждом датафрейме из словаря 'all_raw_data_pr'
all_raw_data_pr = {raw_data[0]: raw_data[1].dropna(how='all', axis=1) for raw_data in all_raw_data_pr.items()}


# In[ ]:


# временно переименуем столбцы
for row_key, row_value in all_raw_data_pr.items():
    row_value.columns = ['0', '1', '2', '3', '4', '5', '6', '7']


# In[ ]:


# создадим список из номеров строк, в которых встречается слово 'Период' в excel файле в столбце '0',
# чтобы в дальнейшем оставить только строки, содержащие таблицу с данными
rows_to_delete_index = []
for raw_data in all_raw_data_pr.values():
    rows_to_delete_index.append(list(raw_data['0']).index('Период'))


# In[ ]:


# объединим списки 'raw_data_name_list_pr' и 'rows_to_delete_index'
# в список кортежей с названием файла и номером строки для удаления
rows_to_delete = list(zip(raw_data_name_list_pr, rows_to_delete_index))


# In[ ]:


# при помощи цикла обновим словарь 'all_raw_data_pr', удалив лишние строки из каждого датафрейма
for raw_data in all_raw_data_pr.items():
    for row in rows_to_delete:
        if raw_data[0] == row[0]:
            all_raw_data_pr.update({raw_data[0]: raw_data[1][row[1]+2:]})


# In[ ]:


# переименуем столбцы
for row_key, row_value in all_raw_data_pr.items():
    row_value.columns = ['Период',
                         'Документ',
                         'Аналитика Дт',
                         'Аналитика Кт',
                         'Счёт Дт',
                         'Сумма Дт',
                         'Счёт Кт',
                         'Сумма Кт']


# In[ ]:


# временно переименуем столбцы
for row_key, row_value in all_raw_data_pr.items():
    if row_value.empty:
        row_value.loc[0] = ['Итого'] + [float('nan')]*(row_value.shape[1]-1)


# In[ ]:


# заменим значения последней ячейки в первом столбце "all_raw_data_pr" на строку "Итого"
for row_key, row_value in all_raw_data_pr.items():
    row_value.iloc[-1, 0] = 'Итого'


# In[ ]:


# создадим список 'source_file_path' и заполним его значениями путей к исходным файлам
source_file_path = []
for name in raw_data_name_list_pr:
    source_file_path.append(os.path.join(source, name))


# In[ ]:


# объединим списки 'raw_data_name_list_pr' и 'source_file_path'
# в список кортежей с названием файла и путём к нему
source_file = list(zip(raw_data_name_list_pr, source_file_path))


# In[ ]:


'''
Цикл, который перезаписывает каждый файл из 'all_raw_data', чтобы сохранить внесённые ранее изменения в его структуру,
а также сохраняет файл в формате xlsx в первоначальную папку
'''
print(f'Перезапись Ведомости амортизации ОС из "{source}"')
for raw_data in tqdm(all_raw_data_pr.items()):
    for file in source_file:
        if raw_data[0] == file[0]:
            raw_data[1].to_excel(file[1], header=True, index=False)


# In[ ]:


# создадим путь к папке 'Отчёты по проводкам'
folder_location_1c = os.path.join(folder_location, 'Отчёты по проводкам')
# создаются только те папки, которые до этого не существовали
if not os.path.exists(folder_location_1c):
    os.makedirs(folder_location_1c)  


# In[ ]:


# создадим excel файл 'Отчёт по проводкам' и добавим в него все датафреймы с исходными данными
with pd.ExcelWriter(os.path.join(source, f'Отчёт по проводкам_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')) as source_file:
    print(f'Создание листов excel из Ведомостей амортизации ОС')
    for raw_data in tqdm(all_raw_data_pr.items()):
        raw_data[1].to_excel(source_file,
                             sheet_name=raw_data[0][raw_data[0].rfind('_')+1:].replace('.xlsx', ''),
                             header=True,
                             index=False)


# In[ ]:


# создадим путь к итоговому excel файлу
book_path_pr = os.path.join(source, f'Отчёт по проводкам_{quarter_name(period_list_end[0][:-5])} {period_list_end[0][-4:]}.xlsx')
# загрузим excel файл для редактирования
book_pr = load_workbook(book_path_pr)


# In[ ]:


columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'] # создадим список столбцов, к которым будем обращаться


# In[ ]:


# создадим таблицы на каждом листе
for sheet in book_pr:
    table = Table(displayName=sheet.title.replace('-','_').replace(' ','_'), ref=f'A1:H{sheet.max_row}') # создадим таблицу
    style = TableStyleInfo(name='TableStyleLight13', # создадим стиль таблицы
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=True)
    table.tableStyleInfo = style # применим стиль
    sheet.add_table(table) # добавим таблицу на лист


# In[ ]:


# цикл для изменения ширины столбцов на каждом листе в excel файле
for sheet in book_pr:
    for column in columns:
        sheet.column_dimensions[column].width = 25


# In[ ]:


# цикл для изменения формата ячеек на каждом листе в excel файле
for sheet in book_pr:
    for column in columns:
        if column in ['A', 'B', 'C', 'D', 'E', 'G']:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[0]
        else:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[3]


# In[ ]:


# цикл для изменения стиля и выравнивания ячеек на каждом листе в excel файле
for sheet in book_pr:
    for column in columns:
        # изменение параметров шрифта
        sheet[f'{column}1'].font = Font(bold=True, 
                                        color='FFFFFFFF')
        # изменение выравнивания
        sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                  vertical='center',
                                                  wrap_text=True)  
    # изменение выравнивания числовых ячеек
    for cell_tuple in sheet[f'E2:E{sheet.max_row}']:
        for cell in cell_tuple:
            cell.alignment = Alignment(horizontal='right',
                                       vertical='top')
    # изменение выравнивания числовых ячеек
    for cell_tuple in sheet[f'H2:H{sheet.max_row}']:
        for cell in cell_tuple:
            cell.alignment = Alignment(horizontal='right',
                                       vertical='top')


# In[ ]:


book_pr.save(book_path_pr) # сохраним excel файл


# In[ ]:


print(f'Перемещение файла "Отчёт по проводкам" из "{source}" в новую папку')
# скопируем итоговый файл с расширением .xlsx в новую папку
shutil.copy(book_path_pr, folder_location)
# удалим итоговый файл с расширением .xlsx из первоначальной папки
os.remove(book_path_pr)

# цикл, который копирует каждый файл из 'raw_data_list' в новую папку и удаляет его из папки c исходными данными
for name in tqdm(raw_data_name_list_pr):
    # удалим исходные данные с расширением .xlsx из первоначальной папки
    os.remove(os.path.join(source, name))
    shutil.copy(os.path.join(source_1c, name), folder_location_1c)

